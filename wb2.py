from openpyxl import Workbook

#new workbook
wb = Workbook()

#adding multiple sheets
ws1 = wb.active

#rename the worksheets
ws1.title = "Sheet1"
ws1['A1'] = 'DATA'
ws1['A2'] = 'SHEET 1'


ws2 = wb.create_sheet(title='Sheet2')
ws2['A1'] = 'DATA'
ws2['A2'] = 'SHEET 2'

ws3 = wb.create_sheet()
ws3['A1'] = 'DATA'
ws3['A2'] = 'SHEET 3'

wb.save("multiple_sheet.xlsx")