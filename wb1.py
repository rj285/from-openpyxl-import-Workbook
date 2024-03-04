from openpyxl import Workbook

wb = Workbook()

ws = wb.active

ws['A1'] = 'hello'
ws['A2'] = 'world'

wb.save('wb1.xlsx')