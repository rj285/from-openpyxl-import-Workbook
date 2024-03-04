from openpyxl import Workbook
from openpyxl.styles import Font,PatternFill

wb = Workbook()

ws1 = wb.active
ws1.title = 'Sheet1'
ws1['a1'] = 'BOLD TXT'
ws1['b1'] = 'ITALIC TXT'
ws1['c1'] = 'UNDERLINED TXT'
ws1['d1'] = 'COLOUR TEST'
ws1['e1'] = 'CELL COLOUR'

bold_font = Font(bold=True)
italic_font = Font(italic=True)
underlined = Font(underline='single')
colors = Font(color='FF0000')
cell_color = PatternFill(start_color='0000FF', end_color= '0000FF', fill_type='solid')

ws1['a1'].font = bold_font
ws1['b1'].font = italic_font
ws1['c1'].font = underlined
ws1['d1'].font = colors
ws1['e1'].font = cell_color

wb.save('wb3.xlsx')
